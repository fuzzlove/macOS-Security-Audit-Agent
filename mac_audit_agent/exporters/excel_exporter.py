from __future__ import annotations

from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from mac_audit_agent.assessment import SecurityAssessment
from mac_audit_agent.exporters.export_models import ExportAssessmentData, ExportOptions, build_export_assessment_data
from mac_audit_agent.ui.severity_styles import get_severity_style, normalize_severity


def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl. Install project dependencies including openpyxl to create .xlsx reports.") from exc
    return Workbook, Alignment, Font, PatternFill, get_column_letter, DataValidation


def _rows_from_dicts(rows: list[dict[str, Any]], columns: list[str]) -> list[list[Any]]:
    if not rows:
        return [["No data was collected for this section."] + [""] * (len(columns) - 1)]
    return [[item.get(column, "") for column in columns] for item in rows]


def _display_header(column: str) -> str:
    return column.replace("_", " ").title().replace("Nist", "NIST").replace("Mitre", "MITRE").replace("Cve", "CVE").replace("Cisa Kev", "CISA KEV")


def _write_sheet(wb, title: str, columns: list[str], rows: list[list[Any]], *, severity_column: str = "severity"):
    _Workbook, Alignment, Font, PatternFill, get_column_letter, _DataValidation = _require_openpyxl()
    ws = wb.create_sheet(title[:31])
    ws.append([_display_header(column) for column in columns])
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    risk_columns = {
        "severity",
        "criticality",
        "priority",
        "impact",
        "risk",
        "risk_level",
        "risk_label",
        "risk_score",
        "cvss",
        "cvss_score",
        "forecast_level",
        "trust",
        "trust_label",
        "trust_status",
        "baseline",
        "baseline_status",
        "finding_status",
        "status",
        "health_status",
    }
    risk_indexes = [index + 1 for index, column in enumerate(columns) if column in risk_columns or column == severity_column]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for risk_index in risk_indexes:
            column_name = columns[risk_index - 1]
            value = row[risk_index - 1].value
            score_type = "cvss" if column_name in {"cvss", "cvss_score"} else "risk_score" if column_name == "risk_score" else None
            normalized = normalize_severity(value, value if score_type else None, score_type)
            colors = get_severity_style(normalized)
            row[risk_index - 1].value = colors.label if value in {None, ""} else value
            row[risk_index - 1].fill = PatternFill("solid", fgColor=colors.background.lstrip("#"))
            row[risk_index - 1].font = Font(bold=True, color=colors.foreground.lstrip("#"))
    for index, column_cells in enumerate(ws.columns, start=1):
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column_cells[:100]) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width
    return ws


def _add_status_dropdown(ws, columns: list[str], row_count: int) -> None:
    if "status" not in columns or row_count <= 0:
        return
    _Workbook, _Alignment, _Font, _PatternFill, get_column_letter, DataValidation = _require_openpyxl()
    status_index = columns.index("status") + 1
    column_letter = get_column_letter(status_index)
    validation = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted Risk,False Positive"',
        allow_blank=True,
    )
    validation.error = "Select a valid remediation status."
    validation.errorTitle = "Invalid Status"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{max(row_count + 1, 2)}")


def _build_workbook(data: ExportAssessmentData):
    Workbook, Alignment, Font, PatternFill, _get_column_letter, _DataValidation = _require_openpyxl()
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    summary_rows = [[key.replace("_", " ").title(), value] for key, value in data.summary.items()]
    summary_rows = [["Hostname", data.metadata.get("hostname", "")], ["macOS Version", data.metadata.get("macos_version", "")], ["Assessment Date", data.metadata.get("assessment_date", "")], *summary_rows]
    _write_sheet(wb, "Executive Summary", ["Field", "Value"], summary_rows, severity_column="")
    findings_columns = ["finding_id", "severity", "confidence", "category", "title", "description", "evidence_summary", "impact", "suggested_fix", "validation_step", "status", "first_seen", "last_seen", "nist_csf", "nist_800_53", "mitre_attack", "cve", "cisa_kev"]
    findings_rows = _rows_from_dicts(data.findings, findings_columns)
    findings_ws = _write_sheet(
        wb,
        "Findings",
        findings_columns,
        findings_rows,
    )
    _add_status_dropdown(findings_ws, findings_columns, len(findings_rows))
    remediation_columns = ["priority", "severity", "finding_id", "recommended_fix", "difficulty", "expected_impact", "validation_step", "owner", "status", "due_date", "notes"]
    remediation_rows = _rows_from_dicts(data.remediation_items, remediation_columns)
    remediation_ws = _write_sheet(wb, "Remediation Plan", remediation_columns, remediation_rows)
    _add_status_dropdown(remediation_ws, remediation_columns, len(remediation_rows))
    _write_sheet(wb, "Apple Exposure", ["advisory", "severity", "affected_component", "current_version", "recommended_action", "update_guidance_title", "update_guidance_summary", "verification_steps", "evidence_preservation_notes", "official_references", "kev_status", "database_checked", "last_successful_update", "freshness_status", "suggested_fix"], _rows_from_dicts(data.apple_exposure, ["advisory", "severity", "affected_component", "current_version", "recommended_action", "update_guidance_title", "update_guidance_summary", "verification_steps", "evidence_preservation_notes", "official_references", "kev_status", "database_checked", "last_successful_update", "freshness_status", "suggested_fix"]))
    _write_sheet(wb, "Network Activity", ["event_id", "severity", "source", "local_address", "local_port", "remote_address", "remote_port", "process", "pid", "signed_status", "evidence", "suggested_fix"], _rows_from_dicts(data.network_activity, ["event_id", "severity", "source", "local_address", "local_port", "remote_address", "remote_port", "process", "pid", "signed_status", "evidence", "suggested_fix"]))
    _write_sheet(wb, "Admin Persistence", ["finding_id", "severity", "type", "path_user", "owner", "permissions", "signed_status", "first_seen", "baseline_status", "evidence", "suggested_fix"], _rows_from_dicts(data.admin_persistence, ["finding_id", "severity", "type", "path_user", "owner", "permissions", "signed_status", "first_seen", "baseline_status", "evidence", "suggested_fix"]))
    _write_sheet(wb, "Physical Devices", ["device_id", "severity", "device_type", "name", "manufacturer", "vendor_id", "product_id", "serial_present", "trust_status", "first_seen", "last_seen", "suggested_fix"], _rows_from_dicts(data.physical_devices, ["device_id", "severity", "device_type", "name", "manufacturer", "vendor_id", "product_id", "serial_present", "trust_status", "first_seen", "last_seen", "suggested_fix"]))
    _write_sheet(wb, "Framework Mapping", ["finding_id", "framework", "control_id", "name", "category", "mapping_confidence", "notes"], _rows_from_dicts(data.framework_mappings, ["finding_id", "framework", "control_id", "name", "category", "mapping_confidence", "notes"]), severity_column="")
    cmmc_summary_rows = [[key, value] for key, value in (data.cmmc_summary or {}).items() if not isinstance(value, (list, dict))]
    _write_sheet(wb, "CMMC Summary", ["field", "value"], cmmc_summary_rows or [["status", "No CMMC readiness data generated."]], severity_column="")
    cmmc_requirement_columns = ["cmmc_id", "level", "domain", "practice_id", "title", "requirement_text", "source_id", "source_version", "implementation_status", "msaa_check_ids", "limitations"]
    cmmc_requirement_rows = []
    for item in data.cmmc_requirements:
        row = dict(item)
        row["msaa_check_ids"] = ", ".join(str(value) for value in row.get("msaa_check_ids", []))
        row["limitations"] = "; ".join(str(value) for value in row.get("limitations", []))
        cmmc_requirement_rows.append(row)
    _write_sheet(wb, "CMMC Requirements", cmmc_requirement_columns, _rows_from_dicts(cmmc_requirement_rows, cmmc_requirement_columns), severity_column="")
    evidence_columns = ["cmmc_level", "cmmc_requirement_id", "domain", "requirement_summary", "related_nist_control", "msaa_check", "evidence_collected", "evidence_location", "evidence_status", "manual_evidence_needed", "suggested_fix", "analyst_notes"]
    _write_sheet(wb, "Evidence Matrix", evidence_columns, _rows_from_dicts(data.cmmc_evidence_matrix, evidence_columns), severity_column="evidence_status")
    poam_columns = ["poam_id", "framework", "requirement_id", "weakness", "risk_level", "source_finding_id", "recommended_fix", "validation_step", "owner", "target_completion_date", "status", "evidence_needed", "notes"]
    poam_ws = _write_sheet(wb, "POA&M", poam_columns, _rows_from_dicts(data.cmmc_poam, poam_columns), severity_column="risk_level")
    _add_status_dropdown(poam_ws, poam_columns, len(data.cmmc_poam))
    source_columns = ["source_id", "framework", "title", "version", "retrieved_at", "source_url", "normative"]
    _write_sheet(wb, "Source Versions", source_columns, _rows_from_dicts(data.cmmc_source_versions, source_columns), severity_column="")
    manual_columns = ["requirement_id", "evidence_needed", "suggested_document_name", "owner", "status", "notes"]
    _write_sheet(wb, "Manual Evidence", manual_columns, _rows_from_dicts(data.cmmc_manual_evidence, manual_columns), severity_column="status")
    _write_sheet(wb, "Events Timeline", ["timestamp", "severity", "event_type", "source", "summary", "related_finding", "evidence", "suggested_fix"], _rows_from_dicts(data.timeline, ["timestamp", "severity", "event_type", "source", "summary", "related_finding", "evidence", "suggested_fix"]))
    _write_sheet(wb, "Visibility Integrity", ["component", "status", "last_success", "last_error", "evidence", "suggested_fix"], _rows_from_dicts(data.visibility_integrity, ["component", "status", "last_success", "last_error", "evidence", "suggested_fix"]), severity_column="")
    application_integrity_columns = ["scope", "status", "manifest_path", "checked_at", "verified_at", "verification_result_id", "manifest_source_type", "current_install_mode", "manifest_app_version", "current_app_version", "manifest_build_id", "current_build_id", "manifest_git_commit", "current_git_commit", "manifest_package_version", "current_package_version", "manifest_root_path", "current_root_path", "manifest_created_at", "manifest_hash", "cached_result", "cache_valid", "cache_invalidated_reason", "ignored_manifests", "matched_count", "mismatched_count", "missing_count", "extra_count", "relative_path", "mismatch_reasons", "exact_mismatch_reason", "recommended_action"]
    _write_sheet(wb, "Application Integrity", application_integrity_columns, _rows_from_dicts(data.application_integrity, application_integrity_columns), severity_column="")
    limitations = [{"limitation": item} for item in data.limitations] or [{"limitation": "No limitations recorded."}]
    _write_sheet(wb, "Limitations", ["limitation"], _rows_from_dicts(limitations, ["limitation"]), severity_column="")
    return wb


def export_assessment_excel(assessment: SecurityAssessment, output_path: Path | None = None, *, options: ExportOptions | None = None) -> Path:
    data = build_export_assessment_data(assessment, options=options)
    if output_path is None:
        host = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in assessment.hostname) or "mac"
        stamp = assessment.created_at.replace(":", "").replace("+", "Z").replace("-", "")[:15]
        from mac_audit_agent.reporting import get_reports_dir

        path = get_reports_dir() / f"MSAA_Security_Assessment_{host}_{stamp}.xlsx"
    else:
        path = output_path
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _build_workbook(data)
    with NamedTemporaryFile(prefix=path.stem, suffix=".tmp.xlsx", dir=path.parent, delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        wb.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return path
